import csv
from typing import Iterator, Dict, List, Tuple


def _read_csv(file_obj) -> Iterator[Dict[str, str]]:
    file_obj.seek(0)
    # Attempt to decode as utf-8; if bytes are passed, wrap accordingly
    try:
        text = file_obj.read().decode('utf-8')
    except AttributeError:
        # Already str
        text = file_obj.read()
    if isinstance(text, bytes):
        text = text.decode('utf-8')
    lines = text.splitlines()
    reader = csv.DictReader(lines)
    for row in reader:
        yield {k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k is not None}


def _read_xlsx(file_obj) -> Iterator[Dict[str, str]]:
    try:
        import openpyxl  # noqa: F401
    except Exception as exc:
        raise RuntimeError('Excel support requires openpyxl. Please install it.') from exc
    file_obj.seek(0)
    from openpyxl import load_workbook
    wb = load_workbook(filename=file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = None
    for idx, row in enumerate(rows):
        if idx == 0:
            headers = [str(h).strip() if h is not None else '' for h in row]
            continue
        if not headers:
            continue
        values = list(row)
        data = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = values[i] if i < len(values) else None
            data[h] = str(val).strip() if isinstance(val, str) else val
        yield data


def detect_and_parse_tabular(file_obj, filename: str) -> Tuple[List[Dict[str, str]], str]:
    name = (filename or '').lower()
    if name.endswith('.csv'):
        rows = list(_read_csv(file_obj))
        return rows, 'csv'
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        rows = list(_read_xlsx(file_obj))
        return rows, 'xlsx'
    # Fallback: try CSV
    rows = list(_read_csv(file_obj))
    return rows, 'csv'


def normalize_customer_row(row: Dict[str, str]) -> Dict[str, object]:
    key_map = {
        'name': ['name', 'full_name', 'customer_name'],
        'email': ['email', 'e-mail'],
        'phone': ['phone', 'mobile', 'contact_number'],
        'company': ['company', 'organization'],
        'address': ['address', 'street'],
        'city': ['city'],
        'state': ['state', 'province'],
        'country': ['country'],
        'zip_code': ['zip_code', 'zipcode', 'postal_code'],
        'is_active': ['is_active', 'active'],
    }
    output: Dict[str, object] = {}
    lowered = {str(k).strip().lower(): v for k, v in row.items()}
    for target, candidates in key_map.items():
        for ck in candidates:
            if ck in lowered:
                val = lowered[ck]
                if target == 'is_active':
                    if isinstance(val, str):
                        output[target] = val.strip().lower() in ('1', 'true', 'yes', 'y')
                    else:
                        output[target] = bool(val)
                else:
                    output[target] = val
                break
    return output


